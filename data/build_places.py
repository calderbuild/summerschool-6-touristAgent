#!/usr/bin/env python3
"""Build the Paris places knowledge base (Excel + CSV) for the touristAgent chatbot.

Data verified 2026-07-23 via official sites + web search for volatile fields
(2026 prices, open/closed status). Stable fields (coordinates, category, district,
nearest transit, accessibility) come from established knowledge; anything not
confirmed is marked so, never guessed. Accessibility should be re-checked on
acceslibre.beta.gouv.fr before the pitch.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    "id",
    "name_en",
    "name_fr",
    "category",
    "arrondissement",
    "latitude",
    "longitude",
    "visit_duration",
    "budget_adult_eur",
    "is_free",
    "opening_hours",
    "nearest_transit",
    "station_step_free",
    "site_wheelchair",
    "accessible_toilet",
    "status",
    "official_url",
    "source",
    "last_verified",
    "notes",
]

ROWS = [
    {
        "id": "eiffel-tower",
        "name_en": "Eiffel Tower",
        "name_fr": "Tour Eiffel",
        "category": "Monument",
        "arrondissement": "7th (75007)",
        "latitude": 48.8584,
        "longitude": 2.2945,
        "visit_duration": "2-3 h",
        "budget_adult_eur": "€36.70 summit (lift); €23.50 2nd floor (lift); €14.80 stairs to 2nd",
        "is_free": "No",
        "opening_hours": "≈09:30-23:45 daily (extended in summer)",
        "nearest_transit": "RER C Champ de Mars-Tour Eiffel; M6 Bir-Hakeim",
        "station_step_free": "Partial (RER C Champ de Mars step-free; M6 Bir-Hakeim not)",
        "site_wheelchair": "Partial (1st & 2nd floors by lift; summit not accessible)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://www.toureiffel.paris/en",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Summit closed to wheelchair users. Book online in advance.",
    },
    {
        "id": "louvre",
        "name_en": "Louvre Museum",
        "name_fr": "Musée du Louvre",
        "category": "Museum",
        "arrondissement": "1st (75001)",
        "latitude": 48.8606,
        "longitude": 2.3376,
        "visit_duration": "3-4 h (half day)",
        "budget_adult_eur": "€22 (EEA residents); €32 (non-EEA)",
        "is_free": "No",
        "opening_hours": "09:00-18:00; late to 21:45 Wed & Fri; closed Tue",
        "nearest_transit": "M1/M7 Palais Royal-Musée du Louvre",
        "station_step_free": "Partial (line 14 Pyramides/Châtelet nearby are step-free)",
        "site_wheelchair": "Yes (fully accessible; free for wheelchair users + 1 companion)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://www.louvre.fr/en",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "New 2026 dual pricing (EEA vs non-EEA). Free under 18, EU 18-25, 1st Sat evening.",
    },
    {
        "id": "notre-dame",
        "name_en": "Notre-Dame Cathedral",
        "name_fr": "Cathédrale Notre-Dame de Paris",
        "category": "Cathedral",
        "arrondissement": "4th (75004)",
        "latitude": 48.8530,
        "longitude": 2.3499,
        "visit_duration": "1 h (cathedral)",
        "budget_adult_eur": "Cathedral free; Bell Towers €16",
        "is_free": "Yes (cathedral)",
        "opening_hours": "≈08:00-19:00 (free reservation online)",
        "nearest_transit": "M4 Cité; RER B/C Saint-Michel-Notre-Dame",
        "station_step_free": "Partial (Cité limited)",
        "site_wheelchair": "Yes (cathedral, step-free entrance); No (towers, ~424 steps)",
        "accessible_toilet": "Unknown (public toilets nearby)",
        "status": "Open",
        "official_url": "https://www.notredamedeparis.fr/en/",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Reopened Dec 2024; towers reopened Sep 2025. Free timed reservation recommended.",
    },
    {
        "id": "arc-de-triomphe",
        "name_en": "Arc de Triomphe",
        "name_fr": "Arc de Triomphe",
        "category": "Monument",
        "arrondissement": "8th (75008)",
        "latitude": 48.8738,
        "longitude": 2.2950,
        "visit_duration": "1 h",
        "budget_adult_eur": "€16 (rooftop); free to view from below",
        "is_free": "No (rooftop)",
        "opening_hours": "10:00-22:30 (seasonal)",
        "nearest_transit": "M1/M2/M6, RER A Charles de Gaulle-Étoile",
        "station_step_free": "Partial",
        "site_wheelchair": "Partial (rooftop ~284 steps; lift for reduced-mobility visitors on request)",
        "accessible_toilet": "Unknown",
        "status": "Open",
        "official_url": "https://www.paris-arc-de-triomphe.fr/en",
        "source": "Official site (CMN) + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Reach the arch via the underground passage (no street crossing). Free under 18, EU 18-25.",
    },
    {
        "id": "musee-orsay",
        "name_en": "Musée d'Orsay",
        "name_fr": "Musée d'Orsay",
        "category": "Museum",
        "arrondissement": "7th (75007)",
        "latitude": 48.8600,
        "longitude": 2.3266,
        "visit_duration": "2-3 h",
        "budget_adult_eur": "€16 online (€14 on-site; €12 Thu evening)",
        "is_free": "No",
        "opening_hours": "09:30-18:00 Tue-Sun; to 21:45 Thu; closed Mon",
        "nearest_transit": "RER C Musée d'Orsay; M12 Solférino",
        "station_step_free": "Partial",
        "site_wheelchair": "Yes (accessible)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://www.musee-orsay.fr/en",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Free 1st Sunday of month; free under 18, EU 18-25. Reservation mandatory.",
    },
    {
        "id": "sacre-coeur",
        "name_en": "Sacré-Cœur Basilica",
        "name_fr": "Basilique du Sacré-Cœur",
        "category": "Basilica",
        "arrondissement": "18th (75018)",
        "latitude": 48.8867,
        "longitude": 2.3431,
        "visit_duration": "1-1.5 h",
        "budget_adult_eur": "Basilica free; dome ≈€8 (confirm)",
        "is_free": "Yes (basilica)",
        "opening_hours": "06:30-22:30 (basilica)",
        "nearest_transit": "M2 Anvers + funicular; M12 Abbesses",
        "station_step_free": "No (hilltop; deep stations with stairs; funicular to summit)",
        "site_wheelchair": "Partial (step-free side access to basilica; hill is difficult; dome not accessible)",
        "accessible_toilet": "Unknown",
        "status": "Open",
        "official_url": "https://www.sacre-coeur-montmartre.com/english/",
        "source": "Official site + knowledge; dome price not re-verified",
        "last_verified": "2026-07-23",
        "notes": "Confirm dome price/access on official site. Funicular covered by a metro ticket.",
    },
    {
        "id": "sainte-chapelle",
        "name_en": "Sainte-Chapelle",
        "name_fr": "Sainte-Chapelle",
        "category": "Monument",
        "arrondissement": "1st (75001)",
        "latitude": 48.8554,
        "longitude": 2.3450,
        "visit_duration": "45 min-1 h",
        "budget_adult_eur": "€12 (combo with Conciergerie €20)",
        "is_free": "No",
        "opening_hours": "09:00-19:00 (seasonal)",
        "nearest_transit": "M4 Cité; RER B/C Saint-Michel",
        "station_step_free": "Partial",
        "site_wheelchair": "Partial (lower chapel step-free; upper chapel via spiral stairs, not accessible)",
        "accessible_toilet": "Unknown",
        "status": "Open",
        "official_url": "https://www.sainte-chapelle.fr/en",
        "source": "Official site (CMN) + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Airport-style security (active courthouse). Upper-chapel stained glass is the highlight.",
    },
    {
        "id": "pantheon",
        "name_en": "Panthéon",
        "name_fr": "Panthéon",
        "category": "Monument",
        "arrondissement": "5th (75005)",
        "latitude": 48.8462,
        "longitude": 2.3464,
        "visit_duration": "1 h",
        "budget_adult_eur": "≈€13 (2025 estimate; confirm 2026)",
        "is_free": "No",
        "opening_hours": "10:00-18:30 (seasonal)",
        "nearest_transit": "RER B Luxembourg; M10 Cardinal Lemoine",
        "station_step_free": "Partial",
        "site_wheelchair": "Partial (nave accessible; dome via stairs, not accessible)",
        "accessible_toilet": "Unknown",
        "status": "Open",
        "official_url": "https://www.paris-pantheon.fr/en",
        "source": "Official site (CMN); 2026 price NOT verified",
        "last_verified": "2026-07-23 (price is a 2025 estimate)",
        "notes": "Price is a 2025 estimate, confirm 2026 on the official site. Crypt of French notables; Foucault pendulum.",
    },
    {
        "id": "jardin-luxembourg",
        "name_en": "Luxembourg Garden",
        "name_fr": "Jardin du Luxembourg",
        "category": "Park",
        "arrondissement": "6th (75006)",
        "latitude": 48.8462,
        "longitude": 2.3372,
        "visit_duration": "1 h",
        "budget_adult_eur": "Free",
        "is_free": "Yes",
        "opening_hours": "≈07:30-21:00 (seasonal, dawn-dusk)",
        "nearest_transit": "RER B Luxembourg; M4 Saint-Sulpice",
        "station_step_free": "Partial",
        "site_wheelchair": "Yes (main paths; some gravel surfaces)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://www.senat.fr/visite/jardin/index.html",
        "source": "Official site + knowledge",
        "last_verified": "2026-07-23",
        "notes": "Free public garden. Gravel paths can be hard for small wheels.",
    },
    {
        "id": "jardin-tuileries",
        "name_en": "Tuileries Garden",
        "name_fr": "Jardin des Tuileries",
        "category": "Park",
        "arrondissement": "1st (75001)",
        "latitude": 48.8635,
        "longitude": 2.3275,
        "visit_duration": "1 h",
        "budget_adult_eur": "Free",
        "is_free": "Yes",
        "opening_hours": "≈07:00-21:00 (seasonal)",
        "nearest_transit": "M1 Tuileries / Concorde",
        "station_step_free": "Partial",
        "site_wheelchair": "Yes (mostly flat; gravel surfaces)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://www.louvre.fr/en/explore/the-palace/the-tuileries-garden",
        "source": "Official site + knowledge",
        "last_verified": "2026-07-23",
        "notes": "Links the Louvre to Place de la Concorde. Gravel paths.",
    },
    {
        "id": "chateau-versailles",
        "name_en": "Palace of Versailles",
        "name_fr": "Château de Versailles",
        "category": "Palace",
        "arrondissement": "Versailles (78000, Yvelines)",
        "latitude": 48.8049,
        "longitude": 2.1204,
        "visit_duration": "half-full day",
        "budget_adult_eur": "Passport €25 (low)-€35 (high); Palace only €21",
        "is_free": "No (palace); gardens free except fountain-show days",
        "opening_hours": "09:00-18:30; closed Mon",
        "nearest_transit": "RER C Versailles Château Rive Gauche",
        "station_step_free": "Partial",
        "site_wheelchair": "Yes (accessible; free for disabled visitors + companion; adapted routes)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://en.chateauversailles.fr",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Outside Paris (~40 min by RER C). EEA reduced rates. Gardens paid on Musical Fountains days.",
    },
    {
        "id": "galeries-lafayette",
        "name_en": "Galeries Lafayette Haussmann",
        "name_fr": "Galeries Lafayette Haussmann",
        "category": "Shopping",
        "arrondissement": "9th (75009)",
        "latitude": 48.8738,
        "longitude": 2.3320,
        "visit_duration": "1-2 h",
        "budget_adult_eur": "Free entry (shopping)",
        "is_free": "Yes (entry)",
        "opening_hours": "10:00-20:00; Sun 11:00-20:00",
        "nearest_transit": "M7/M9 Chaussée d'Antin-La Fayette; RER A Auber",
        "station_step_free": "Partial",
        "site_wheelchair": "Yes (department store with lifts)",
        "accessible_toilet": "Yes",
        "status": "Open",
        "official_url": "https://haussmann.galerieslafayette.com/en/",
        "source": "Official site + knowledge",
        "last_verified": "2026-07-23",
        "notes": "Belle Époque glass dome; free rooftop terrace with city views.",
    },
    {
        "id": "centre-pompidou",
        "name_en": "Centre Pompidou",
        "name_fr": "Centre Pompidou",
        "category": "Museum",
        "arrondissement": "4th (75004)",
        "latitude": 48.8607,
        "longitude": 2.3522,
        "visit_duration": "N/A",
        "budget_adult_eur": "N/A (closed)",
        "is_free": "N/A",
        "opening_hours": "Closed",
        "nearest_transit": "M11 Rambuteau; M1/4/7/11/14 Châtelet",
        "station_step_free": "Partial",
        "site_wheelchair": "N/A (closed)",
        "accessible_toilet": "N/A",
        "status": "CLOSED for renovation (22 Sep 2025 - 2030)",
        "official_url": "https://www.centrepompidou.fr/en/",
        "source": "Official site + web search 2026-07-23",
        "last_verified": "2026-07-23",
        "notes": "Closed for a 5-year renovation until 2030. Collection shown at partner venues (Constellation programme). Do NOT recommend visiting.",
    },
]

OUT_DIR = "/Users/calder/epsi-paris-team6/data"
os.makedirs(OUT_DIR, exist_ok=True)

# ---- CSV (easy import into a database / no-code table) ----
csv_path = os.path.join(OUT_DIR, "paris-places.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=COLUMNS)
    w.writeheader()
    for r in ROWS:
        w.writerow(r)

# ---- Excel (formatted knowledge base) ----
wb = Workbook()
ws = wb.active
ws.title = "Places"

header_fill = PatternFill("solid", fgColor="12263A")
header_font = Font(bold=True, color="FFFFFF", size=11)
wrap_top = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9DEE5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append([c.replace("_", " ") for c in COLUMNS])
for r in ROWS:
    ws.append([r[c] for c in COLUMNS])

# header style
for col in range(1, len(COLUMNS) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)

# body style + zebra
zebra = PatternFill("solid", fgColor="F5F7FA")
for row in range(2, len(ROWS) + 2):
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = wrap_top
        cell.border = border
        if row % 2 == 0:
            cell.fill = zebra
    # flag the closed row in red
    status_cell = ws.cell(row=row, column=COLUMNS.index("status") + 1)
    if str(status_cell.value).startswith("CLOSED"):
        status_cell.font = Font(bold=True, color="B00020")

widths = {
    "id": 18,
    "name_en": 24,
    "name_fr": 26,
    "category": 12,
    "arrondissement": 22,
    "latitude": 10,
    "longitude": 10,
    "visit_duration": 14,
    "budget_adult_eur": 40,
    "is_free": 16,
    "opening_hours": 34,
    "nearest_transit": 34,
    "station_step_free": 34,
    "site_wheelchair": 40,
    "accessible_toilet": 20,
    "status": 24,
    "official_url": 40,
    "source": 30,
    "last_verified": 26,
    "notes": 46,
}
for i, c in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths[c]

ws.freeze_panes = "C2"  # freeze header row + id/name_en
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(ROWS) + 1}"

# ---- Schema / README sheet ----
doc = wb.create_sheet("README")
doc.column_dimensions["A"].width = 24
doc.column_dimensions["B"].width = 96
notes = [
    (
        "Paris places knowledge base",
        "Grounding data for the touristAgent chatbot (Team 6).",
    ),
    (
        "Rows",
        f"{len(ROWS)} places ({sum(1 for r in ROWS if not str(r['status']).startswith('CLOSED'))} open + 1 closed for renovation). Meets the MSPR minimum of 10 referenced places.",
    ),
    (
        "Verified",
        "2026-07-23. Prices and open/closed status checked against official sites via web search; stable fields (coordinates, category, transit) from established knowledge.",
    ),
    (
        "Honesty rule",
        "Unknown values are written as 'Unknown' or flagged, never guessed. Panth‚on price and Sacr‚-C≈ìur dome price are estimates and are marked as such.",
    ),
    (
        "Accessibility",
        "step-free / wheelchair / toilet fields are a starting point. Re-check each site on acceslibre.beta.gouv.fr and RATP/IDFM for live station lift status before the pitch.",
    ),
    (
        "Category",
        "Monument, Museum, Cathedral, Basilica, Park, Palace, Shopping. Extend with Food / Transport / Health / Emergency to cover all MSPR categories.",
    ),
    (
        "budget_adult_eur",
        "Visitor cost for one adult (entry). 'Free' where entry is free. Reduced rates (under 18, EU 18-25, first-Sunday, etc.) are in notes, not modelled as columns yet.",
    ),
    ("Coordinates", "WGS84 lat/lng for the interactive map. Accurate to ~4 decimals."),
    (
        "How to extend",
        "Add one row per place; keep the same columns. Import paris-places.csv into your database / no-code table, or read the .xlsx directly.",
    ),
    (
        "Closed places",
        "Centre Pompidou is CLOSED for renovation until 2030. Keep it in the base with status=CLOSED so the chatbot knows not to recommend it.",
    ),
]
doc.append(["Field", "Meaning"])
for k, v in notes:
    doc.append([k, v])
for col in range(1, 3):
    c = doc.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
for row in range(1, len(notes) + 2):
    for col in range(1, 3):
        doc.cell(row=row, column=col).alignment = wrap_top

xlsx_path = os.path.join(OUT_DIR, "paris-places.xlsx")
wb.save(xlsx_path)

# ---- TypeScript module: the app's grounding knowledge base ----
import json as _json


def _ts(s):
    return _json.dumps(s, ensure_ascii=False)


lib_ts_path = "/Users/calder/epsi-paris-team6/mvp-demo/lib/places.ts"
L = []
L.append(
    "// Curated Paris places knowledge base for the touristAgent chatbot grounding."
)
L.append(
    "// Generated from data/build_places.py; human-facing copy: data/paris-places.xlsx."
)
L.append(
    "// Verified 2026-07-23 (2026 prices + open/closed status). Unknowns are honest, not guessed."
)
L.append("")
L.append("export type PlaceCategory =")
L.append('  | "Monument"')
L.append('  | "Museum"')
L.append('  | "Cathedral"')
L.append('  | "Basilica"')
L.append('  | "Park"')
L.append('  | "Palace"')
L.append('  | "Shopping";')
L.append("")
L.append("export interface Place {")
L.append("  id: string;")
L.append("  nameEn: string;")
L.append("  nameFr: string;")
L.append("  category: PlaceCategory;")
L.append("  arrondissement: string;")
L.append("  coord: { lat: number; lng: number };")
L.append("  visitDuration: string;")
L.append('  budget: string; // adult entry cost in EUR, or "Free"')
L.append("  free: boolean; // entry to the site is free")
L.append("  openingHours: string;")
L.append("  nearestTransit: string;")
L.append("  stationStepFree: string;")
L.append("  wheelchair: string;")
L.append("  accessibleToilet: string;")
L.append('  status: "open" | "closed";')
L.append("  officialUrl: string;")
L.append("  source: string;")
L.append("  lastVerified: string;")
L.append("  notes: string;")
L.append("}")
L.append("")
L.append("export const PLACES: Place[] = [")
for r in ROWS:
    free = str(r["is_free"]).strip().lower().startswith("yes")
    status = "closed" if str(r["status"]).startswith("CLOSED") else "open"
    L.append("  {")
    L.append(f"    id: {_ts(r['id'])},")
    L.append(f"    nameEn: {_ts(r['name_en'])},")
    L.append(f"    nameFr: {_ts(r['name_fr'])},")
    L.append(f"    category: {_ts(r['category'])},")
    L.append(f"    arrondissement: {_ts(r['arrondissement'])},")
    L.append(f"    coord: {{ lat: {r['latitude']}, lng: {r['longitude']} }},")
    L.append(f"    visitDuration: {_ts(r['visit_duration'])},")
    L.append(f"    budget: {_ts(r['budget_adult_eur'])},")
    L.append(f"    free: {'true' if free else 'false'},")
    L.append(f"    openingHours: {_ts(r['opening_hours'])},")
    L.append(f"    nearestTransit: {_ts(r['nearest_transit'])},")
    L.append(f"    stationStepFree: {_ts(r['station_step_free'])},")
    L.append(f"    wheelchair: {_ts(r['site_wheelchair'])},")
    L.append(f"    accessibleToilet: {_ts(r['accessible_toilet'])},")
    L.append(f"    status: {_ts(status)},")
    L.append(f"    officialUrl: {_ts(r['official_url'])},")
    L.append(f"    source: {_ts(r['source'])},")
    L.append(f"    lastVerified: {_ts(r['last_verified'])},")
    L.append(f"    notes: {_ts(r['notes'])},")
    L.append("  },")
L.append("];")
L.append("")
with open(lib_ts_path, "w", encoding="utf-8") as f:
    f.write("\n".join(L))

print("Wrote:")
print(" ", xlsx_path)
print(" ", csv_path)
print(" ", lib_ts_path)
print("Rows:", len(ROWS), "| Columns:", len(COLUMNS))
print(
    "Open:",
    sum(1 for r in ROWS if not str(r["status"]).startswith("CLOSED")),
    "| Closed:",
    sum(1 for r in ROWS if str(r["status"]).startswith("CLOSED")),
)
